from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, numbers
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter  
from io import BytesIO
import urllib.request
from datetime import datetime, date

from ..models.grades import Children,Grades

class ExcelReport:
    #wb = load_workbook(filename = BytesIO(file),data_only=True)
    wb = load_workbook('C:/Users/danie/PycharmProjects/school_app/report_format.xlsx')
    ws = wb.active
    #ws.merge_cells()
    student_name_box = 'B2'
    student_lastname_box = 'B3'
    teacher_box = 'E3'
    date_of_report_box = 'B4'
    class_name_box = 'E2'
    bimester_box = 'E4'


    first_grade_row = 8
    __grades_length = 0
    


    def __init__(self,student,subject):

        self.student_name = student.name
        self.student_lastname = student.lastname
        self.student_id = student.child_id
        self.teacher = f'{subject.teacher.name} {subject.teacher.lastname}'
        self.subject = f'{subject.grade_group.name} {subject.subject.name}'
        self.subject_id = subject.grade_subject_id
        self.bimester = '1'
        self.date = date.today()

    def average(self):

        grades = Grades.by_class_student(self.student_id,self.subject_id)
        
        current_grades = [grade for grade in grades if not grade.grade == None]
        self.grades_length = len(current_grades)

        average_grade_row = self.first_grade_row + self.grades_length + 1
        performance_message_row = self.first_grade_row + self.grades_length + 3
        comparison_title_row = self.first_grade_row + self.grades_length + 6
        comparison_message_row = self.first_grade_row + self.grades_length + 7

        for row in range(self.first_grade_row, self.first_grade_row + self.grades_length):
            
            self.ws.merge_cells(start_row = row,end_row = row, start_column =  1, end_column = 2)
            self.ws.merge_cells(start_row = row,end_row = row, start_column =  3, end_column = 4)
            self.ws.merge_cells(start_row = row,end_row = row, start_column =  5, end_column = 6)
            
            self.ws[f'A{str(row)}'] = current_grades[row - self.first_grade_row].event.name
            self.ws[f'C{str(row)}'] = current_grades[row - self.first_grade_row].event.date
            self.ws[f'E{str(row)}'] = current_grades[row - self.first_grade_row].grade

        current_grades_grade = [grade.grade for grade in current_grades]
        average_grade = (sum(current_grades_grade)) / self.grades_length

        if average_grade < 60:
            permormance = 'POOR'
        elif average_grade > 60 and average_grade < 70:
            permormance = 'LOW'
        elif average_grade > 70 and average_grade < 80:
            permormance = 'ACCEPTABLE'
        elif average_grade > 80 and average_grade < 90:
            permormance = 'GOOD'
        else:
            permormance = 'EXCELLENT' 

        performance_message = ("The current average of the "
            f"student ({average_grade}/100) corresponds to a {permormance} performance")  
        
        
        self.ws.merge_cells(start_row = average_grade_row,
            end_row = average_grade_row,start_column =  1, end_column = 4)
        
        self.ws.merge_cells(start_row = average_grade_row,
            end_row = average_grade_row,start_column =  5, end_column = 6)
        
        self.ws[f'A{str(average_grade_row)}'] = 'CURRENT AVERAGE:'
        self.ws[f'E{str(average_grade_row)}'] = average_grade

        self.ws.merge_cells(start_row = performance_message_row,
            end_row = performance_message_row,start_column =  1, end_column = 6)
        
        self.ws[f'A{str(performance_message_row)}'] = performance_message
        

    def generate_report(self):

        self.ws[self.student_name_box] = self.student_name
        self.ws[self.student_lastname_box] = self.student_lastname
        self.ws[self.date_of_report_box] = self.date
        self.ws[self.class_name_box] = self.subject
        self.ws[self.teacher_box] = self.teacher
        self.ws[self.bimester_box] = self.bimester

        self.wb.save('fucking_report.xlsx')
            
    # notes = [88,85.3,75,77.6,79.34,81.43,88.5,92,92.6,96.24,83,85.6,80.04,
    #     74,81.4,83.56,90.04,86.87,79.8,88.56]
    # print(len(notes))

    # notes = sorted(notes)
    # print(notes)

    # percentile = ((notes.index(81.4) + 1)/len(notes)) * 100

    # print(percentile)